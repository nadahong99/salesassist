from pathlib import Path
from datetime import datetime
import pandas as pd

OUTPUT_DIR = Path(__file__).parent.parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


def _margin(sell_price: float, purchase_price: float, commission_pct: float,
            shipping_cost: float, other_cost: float, tax_rate: float = 0.0) -> float:
    """마진율(%) 계산. tax_rate > 0 이면 판매가 기준 부가세를 비용으로 차감."""
    if sell_price <= 0:
        return 0.0
    commission = sell_price * commission_pct / 100
    tax = sell_price * tax_rate / 100
    profit = sell_price - purchase_price - commission - shipping_cost - other_cost - tax
    return round(profit / sell_price * 100, 2)


def _build_df(items: list, config: dict, platform: str) -> pd.DataFrame:
    commission_key = f"{platform}_commission"
    commission_pct = config.get(commission_key, 0)
    shipping_cost = config.get("shipping_cost", 0)
    other_cost = config.get("other_cost", 0)
    tax_rate = config.get("tax_rate", 0.0)
    price_key = f"{platform}_price"

    rows = []
    for item in items:
        sell_price = item.get(price_key) or 0
        purchase_price = item.get("purchase_price", 0)
        commission_amt = round(sell_price * commission_pct / 100, 0)
        tax_amt = round(sell_price * tax_rate / 100, 0)
        margin = _margin(sell_price, purchase_price, commission_pct, shipping_cost, other_cost, tax_rate)
        rows.append({
            "상품명": item.get("name", ""),
            "소싱처": item.get("source", ""),
            "소싱 URL": item.get("url", ""),
            "매입가": purchase_price,
            "판매가": sell_price,
            f"수수료({commission_pct}%)": commission_amt,
            "배송비": shipping_cost,
            "기타비용": other_cost,
            f"부가세({tax_rate}%)": tax_amt,
            "마진(원)": round(sell_price - purchase_price - commission_amt - shipping_cost - other_cost - tax_amt, 0),
            "마진율(%)": margin,
            "메모": item.get("memo", ""),
            "반품 발생": "Y" if item.get("return_occurred", False) else "",
            "환불 완료": "Y" if item.get("return_checklist", {}).get("refund_complete", False) else "",
            "고객 연락": "Y" if item.get("return_checklist", {}).get("customer_contact", False) else "",
            "재고 반영": "Y" if item.get("return_checklist", {}).get("stock_updated", False) else "",
            "배송비 처리": "Y" if item.get("return_checklist", {}).get("shipping_fee_handled", False) else "",
        })
    return pd.DataFrame(rows)


def _save_excel(df: pd.DataFrame, filename: str, target_margin: float) -> str:
    path = OUTPUT_DIR / filename
    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="상품목록")
        ws = writer.sheets["상품목록"]
        from openpyxl.styles import PatternFill, Font
        green_fill = PatternFill("solid", fgColor="C6EFCE")
        yellow_fill = PatternFill("solid", fgColor="FFEB9C")
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        header_font = Font(bold=True)
        margin_col_idx = df.columns.get_loc("마진율(%)")  # 0-based index within the row tuple

        for col_cell in ws[1]:
            col_cell.font = header_font

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            margin_cell = row[margin_col_idx]
            try:
                val = float(margin_cell.value or 0)
            except (ValueError, TypeError):
                val = 0.0
            if val >= target_margin:
                margin_cell.fill = green_fill
            elif val >= target_margin * 0.5:
                margin_cell.fill = yellow_fill
            else:
                margin_cell.fill = red_fill

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)

    return str(path)


def export_naver(items: list, config: dict) -> str:
    df = _build_df(items, config, "naver")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return _save_excel(df, f"naver_{ts}.xlsx", config.get("target_margin", 20))


def export_coupang(items: list, config: dict) -> str:
    df = _build_df(items, config, "coupang")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return _save_excel(df, f"coupang_{ts}.xlsx", config.get("target_margin", 20))
